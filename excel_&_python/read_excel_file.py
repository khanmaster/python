# Python program to fetch data from an excel file

# import openpyxl module
import openpyxl

# Give the location of the file
path = "dinner.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active
max_col = sheet_obj.max_column
max_row = sheet_obj.max_row
# Cell objects also have row, column,
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or
# column integer is 1, not 0.

# Cell object is created by using
# sheet object's cell() method.
cell_obj = sheet_obj.cell(row=2, column=2)
# Print value of cell object
# using the value attribute
print("The 3 course meal costs £", cell_obj.value)

# Exercise
# Determine and print total number of rows and columns from the excel file
# print the names and cost of the meals
for i in range(1, max_col + 1):
    cell_obj = sheet_obj.cell(row = 1, column = i)
    print(cell_obj.value)

for i in range (1, max_row + 1):
    cell_obj = sheet_obj.cell(row = 2, column = i)
    print(cell_obj.value, end = " ")



# print(sheet_obj.max_row)
# print(sheet_obj.max_column)

# Excersise
